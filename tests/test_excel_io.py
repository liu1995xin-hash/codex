from pathlib import Path
import stat

import pytest
from openpyxl import Workbook, load_workbook

from extract_expert_ids.excel_io import (
    ExcelIOError,
    default_output_path,
    list_sheet_names,
    next_result_sheet_name,
    read_sheet_values,
    write_result_workbook,
)


def make_sample_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    rows = [
        ["空白", "", ""],
        ["空白", "2024147", "2024132"],
        ["日期", "1", "2"],
        ["", "刘哥", "八宝饭"],
    ]
    for row in rows:
        ws.append(row)
    ws["D1"] = "=SUM(1,2)"
    wb.save(path)


def test_list_sheet_names_for_xlsx(tmp_path):
    path = tmp_path / "数据源.xlsx"
    make_sample_xlsx(path)

    assert list_sheet_names(path) == ["Sheet1"]


def test_read_sheet_values_for_xlsx(tmp_path):
    path = tmp_path / "数据源.xlsx"
    make_sample_xlsx(path)

    rows = read_sheet_values(path, "Sheet1")

    assert rows[1][1] == "2024147"
    assert rows[3][1] == "刘哥"


def test_default_output_path_uses_combination_suffix(tmp_path):
    input_path = tmp_path / "数据源.xlsx"

    result = default_output_path(input_path)

    assert result.name == "数据源_组合筛选结果.xlsx"


def test_next_result_sheet_name_increments():
    assert next_result_sheet_name(["Sheet1"]) == "结果"
    assert next_result_sheet_name(["Sheet1", "结果"]) == "结果1"
    assert next_result_sheet_name(["结果", "结果1"]) == "结果2"


def test_write_result_workbook_copies_xlsx_and_adds_result_sheet(tmp_path):
    input_path = tmp_path / "输入.xlsx"
    output_path = tmp_path / "输出.xlsx"
    make_sample_xlsx(input_path)

    write_result_workbook(
        input_path,
        output_path,
        [
            ["专家ID组合", "组合人数", "出现期次数", "出现期次"],
            ["A，B，C，D，E", 5, 4, "2024001，2024002"],
        ],
    )

    wb = load_workbook(output_path, data_only=False)
    assert "Sheet1" in wb.sheetnames
    assert "结果" in wb.sheetnames
    assert wb["Sheet1"]["D1"].value == "=SUM(1,2)"
    assert wb["结果"]["A1"].value == "专家ID组合"
    assert wb["结果"]["B2"].value == 5


def test_write_result_workbook_increments_existing_result_sheet(tmp_path):
    input_path = tmp_path / "输入.xlsx"
    output_path = tmp_path / "输出.xlsx"
    wb = Workbook()
    wb.active.title = "结果"
    wb.create_sheet("结果1")
    wb.save(input_path)

    write_result_workbook(input_path, output_path, [["专家ID组合"]])

    output_wb = load_workbook(output_path)
    assert "结果2" in output_wb.sheetnames


def test_write_result_workbook_rejects_same_input_and_output_path(tmp_path):
    input_path = tmp_path / "输入.xlsx"
    make_sample_xlsx(input_path)

    with pytest.raises(ExcelIOError, match="输出路径不能与输入路径相同"):
        write_result_workbook(input_path, input_path, [["专家ID组合"]])


def test_write_result_workbook_handles_readonly_xlsx_input(tmp_path):
    input_path = tmp_path / "只读输入.xlsx"
    output_path = tmp_path / "输出.xlsx"
    make_sample_xlsx(input_path)
    input_path.chmod(stat.S_IREAD)

    try:
        write_result_workbook(input_path, output_path, [["专家ID组合"], ["A，B，C，D，E"]])
    finally:
        input_path.chmod(stat.S_IREAD | stat.S_IWRITE)

    wb = load_workbook(output_path)
    assert "结果" in wb.sheetnames
    assert output_path.stat().st_mode & stat.S_IWRITE
