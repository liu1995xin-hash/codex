from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from number_based_extractor.excel_io import (
    ExcelWorkflowError,
    read_config_workbook,
    read_sheet_rows,
    scan_period_folder,
    write_result_workbook,
    write_template_workbook,
)
from number_based_extractor.models import PeriodFile


def make_period_workbook(
    path: Path,
    *,
    sheet_name: str = "提取结果",
    period: str = "2026-001期",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["期号", "专家名称", "玩法", "本期推荐"])
    ws.append([period, "A", "蓝球定", "01,03"])
    wb.save(path)


def test_scan_period_folder_reads_non_temp_xlsx_and_prefers_result_sheet(tmp_path):
    make_period_workbook(tmp_path / "2026033双色球专家提取结果.xlsx", period="033期")
    make_period_workbook(tmp_path / "~$2026034.xlsx", period="034期")
    (tmp_path / "ignore.txt").write_text("x", encoding="utf-8")

    periods, logs = scan_period_folder(tmp_path)

    assert len(periods) == 1
    assert periods[0].display_period == "033期"
    assert periods[0].standard_period == "2026033"
    assert periods[0].short_period == "033"
    assert periods[0].sheet_name == "提取结果"
    assert any("扫描" in line for line in logs)


def test_read_sheet_rows_falls_back_to_first_sheet_and_validates_headers(tmp_path):
    path = tmp_path / "2026001.xlsx"
    make_period_workbook(path, sheet_name="Sheet1")

    rows = read_sheet_rows(path)

    assert rows[0] == ["期号", "专家名称", "玩法", "本期推荐"]
    assert rows[1][1] == "A"


def test_read_sheet_rows_rejects_invalid_first_four_headers(tmp_path):
    path = tmp_path / "bad.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["期号", "专家名称", "错误", "本期推荐"])
    wb.save(path)

    with pytest.raises(ExcelWorkflowError, match="表头"):
        read_sheet_rows(path)


def test_write_template_workbook_uses_required_columns(tmp_path):
    path = tmp_path / "模板.xlsx"
    periods = [
        PeriodFile(
            file_path=tmp_path / "2026033.xlsx",
            file_name="2026033.xlsx",
            sheet_name="提取结果",
            display_period="033期",
            standard_period="2026033",
            short_period="033",
            selected_number=8,
        )
    ]

    write_template_workbook(path, periods)

    wb = load_workbook(path)
    ws = wb.active
    assert ws.title == "号码配置"
    assert [ws.cell(1, col).value for col in range(1, 5)] == ["期号", "标准期号", "文件名", "筛选号码"]
    assert [ws.cell(2, col).value for col in range(1, 5)] == ["033期", "2026033", "2026033.xlsx", 8]


def test_read_config_workbook_supports_standard_period_and_filter_number(tmp_path):
    path = tmp_path / "配置.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["期号", "标准期号", "文件名", "筛选号码"])
    ws.append(["033期", "2026033", "2026033.xlsx", "01"])
    ws.append(["034期", "", "2026034.xlsx", "20"])
    wb.save(path)

    updates = read_config_workbook(path)

    assert [(u.key, u.number, u.row_index) for u in updates] == [
        ("2026033", 1, 2),
        ("034期", None, 3),
    ]


def test_read_config_workbook_prefers_filter_number_over_generic_number_header(tmp_path):
    path = tmp_path / "配置.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["标准期号", "号码", "筛选号码"])
    ws.append(["2026001", "3", "5"])
    wb.save(path)

    updates = read_config_workbook(path)

    assert [(u.key, u.number) for u in updates] == [("2026001", 5)]


def test_write_result_workbook_outputs_filter_result_sheet_and_header(tmp_path):
    path = tmp_path / "结果.xlsx"

    write_result_workbook(path, [["033期", "A", "蓝球定", "01", 1]])

    wb = load_workbook(path)
    ws = wb["筛选结果"]
    assert [ws.cell(1, col).value for col in range(1, 6)] == ["期号", "专家名称", "玩法", "本期推荐", "筛选号码"]
    assert [ws.cell(2, col).value for col in range(1, 6)] == ["033期", "A", "蓝球定", "01", 1]
