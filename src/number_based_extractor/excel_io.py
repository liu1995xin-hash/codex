from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook

from .core import INPUT_HEADER, RESULT_HEADER, build_period_identity, normalize_cell, normalize_filter_number, validate_input_header
from .models import ConfigUpdate, PeriodFile


class ExcelWorkflowError(Exception):
    """Raised when number-based Excel workflow input or output is invalid."""


TEMPLATE_HEADER = ["期号", "标准期号", "文件名", "筛选号码"]
CONFIG_KEY_HEADERS = ["标准期号", "期号"]
CONFIG_NUMBER_HEADERS = ["筛选号码", "号码"]


def _ensure_xlsx_path(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ExcelWorkflowError("文件扩展名必须是 .xlsx")


def _trim_trailing_empty(rows: list[list[object]]) -> list[list[object]]:
    while rows and all(value in (None, "") for value in rows[-1]):
        rows.pop()
    width = max((len(row) for row in rows), default=0)
    while width > 0:
        if any(len(row) >= width and row[width - 1] not in (None, "") for row in rows):
            break
        width -= 1
    return [row[:width] for row in rows]


def _load_workbook_readonly(path: Path):
    _ensure_xlsx_path(path)
    if not path.exists():
        raise ExcelWorkflowError(f"文件不存在：{path}")
    if not path.is_file():
        raise ExcelWorkflowError(f"路径不是文件：{path}")
    try:
        return load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelWorkflowError(f"无法读取 Excel 文件：{path}，原因：{exc}") from exc


def _select_data_sheet_name(workbook) -> str:
    if "提取结果" in workbook.sheetnames:
        return "提取结果"
    if workbook.sheetnames:
        return workbook.sheetnames[0]
    raise ExcelWorkflowError("Excel 文件没有可读取的 sheet")


def read_sheet_rows(path: str | Path) -> list[list[object]]:
    workbook_path = Path(path)
    wb = _load_workbook_readonly(workbook_path)
    try:
        sheet_name = _select_data_sheet_name(wb)
        ws = wb[sheet_name]
        rows = _trim_trailing_empty([list(row) for row in ws.iter_rows(values_only=True)])
    finally:
        wb.close()

    try:
        validate_input_header(rows)
    except Exception as exc:
        raise ExcelWorkflowError(f"{workbook_path.name} 表头不符合要求：{exc}") from exc
    return rows


def _first_display_period(rows: Sequence[Sequence[object]]) -> str:
    for row in rows[1:]:
        if row:
            period = normalize_cell(row[0])
            if period:
                return period
    return ""


def scan_period_folder(folder: str | Path) -> tuple[list[PeriodFile], list[str]]:
    folder_path = Path(folder)
    if not folder_path.exists():
        raise ExcelWorkflowError(f"文件夹不存在：{folder_path}")
    if not folder_path.is_dir():
        raise ExcelWorkflowError(f"路径不是文件夹：{folder_path}")

    logs: list[str] = []
    periods: list[PeriodFile] = []
    files = sorted(
        path
        for path in folder_path.iterdir()
        if path.is_file()
        and path.suffix.lower() == ".xlsx"
        and not path.name.startswith("~$")
    )
    for path in files:
        try:
            wb = _load_workbook_readonly(path)
            try:
                sheet_name = _select_data_sheet_name(wb)
                ws = wb[sheet_name]
                rows = _trim_trailing_empty([list(row) for row in ws.iter_rows(values_only=True)])
            finally:
                wb.close()
            validate_input_header(rows)
            identity = build_period_identity(path, _first_display_period(rows))
            periods.append(
                PeriodFile(
                    file_path=path,
                    file_name=path.name,
                    sheet_name=sheet_name,
                    display_period=identity.display_period,
                    standard_period=identity.standard_period,
                    short_period=identity.short_period,
                )
            )
            logs.append(f"[扫描] {path.name} -> {identity.display_period} / {identity.standard_period}")
        except Exception as exc:
            logs.append(f"[异常] {path.name} 读取失败：{exc}")
    return periods, logs


def write_template_workbook(path: str | Path, periods: Sequence[PeriodFile]) -> None:
    output_path = Path(path)
    _ensure_xlsx_path(output_path)
    if not output_path.parent.exists():
        raise ExcelWorkflowError(f"输出目录不存在：{output_path.parent}")

    wb = Workbook()
    ws = wb.active
    ws.title = "号码配置"
    ws.append(TEMPLATE_HEADER)
    for period in periods:
        ws.append(
            [
                period.display_period,
                period.standard_period,
                period.file_name,
                period.selected_number,
            ]
        )
    try:
        wb.save(output_path)
    except PermissionError as exc:
        raise ExcelWorkflowError(f"输出文件写入失败，可能被占用：{output_path}") from exc
    except Exception as exc:
        raise ExcelWorkflowError(f"输出文件写入失败：{exc}") from exc


def _header_map(row: Sequence[object]) -> dict[str, int]:
    return {
        normalize_cell(value): index
        for index, value in enumerate(row)
        if normalize_cell(value)
    }


def _find_first_header_index(headers: dict[str, int], names: Sequence[str]) -> int | None:
    for name in names:
        if name in headers:
            return headers[name]
    return None


def read_config_workbook(path: str | Path) -> list[ConfigUpdate]:
    workbook_path = Path(path)
    wb = _load_workbook_readonly(workbook_path)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = _trim_trailing_empty([list(row) for row in ws.iter_rows(values_only=True)])
    finally:
        wb.close()

    if not rows:
        return []

    headers = _header_map(rows[0])
    standard_index = headers.get("标准期号")
    display_index = headers.get("期号")
    number_index = _find_first_header_index(headers, CONFIG_NUMBER_HEADERS)

    if number_index is None or (standard_index is None and display_index is None):
        key_index = 0
        number_index = 1
        start_index = 0
    else:
        key_index = standard_index if standard_index is not None else display_index
        start_index = 1

    updates: list[ConfigUpdate] = []
    for row_number, row in enumerate(rows[start_index:], start=start_index + 1):
        if len(row) <= max(key_index, number_index):
            continue
        key = normalize_cell(row[key_index])
        if not key and standard_index is not None and display_index is not None and display_index < len(row):
            key = normalize_cell(row[display_index])
        if not key:
            continue
        updates.append(
            ConfigUpdate(
                key=key,
                number=normalize_filter_number(row[number_index]),
                source_label=str(workbook_path),
                row_index=row_number,
            )
        )
    return updates


def write_result_workbook(path: str | Path, rows: Sequence[Sequence[object]]) -> None:
    output_path = Path(path)
    _ensure_xlsx_path(output_path)
    if not output_path.parent.exists():
        raise ExcelWorkflowError(f"输出目录不存在：{output_path.parent}")

    wb = Workbook()
    ws = wb.active
    ws.title = "筛选结果"
    ws.append(RESULT_HEADER)
    for row in rows:
        ws.append(list(row))
    try:
        wb.save(output_path)
    except PermissionError as exc:
        raise ExcelWorkflowError(f"输出文件写入失败，可能被占用：{output_path}") from exc
    except Exception as exc:
        raise ExcelWorkflowError(f"输出文件写入失败：{exc}") from exc
