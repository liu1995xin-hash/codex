from __future__ import annotations

import shutil
import stat
from pathlib import Path
from typing import Sequence

import openpyxl
import xlrd
from openpyxl import Workbook


SUPPORTED_INPUT_SUFFIXES = {".xlsx", ".xls"}


class ExcelIOError(Exception):
    """Raised when workbook reading or writing fails."""


def ensure_supported_input(path: Path) -> None:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_INPUT_SUFFIXES:
        raise ExcelIOError("输入文件扩展名必须是 .xlsx 或 .xls")
    if not path.exists():
        raise ExcelIOError(f"输入文件不存在：{path}")
    if not path.is_file():
        raise ExcelIOError(f"输入路径不是文件：{path}")


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_组合筛选结果.xlsx")


def next_result_sheet_name(existing_names: Sequence[str]) -> str:
    names = set(existing_names)
    if "结果" not in names:
        return "结果"
    index = 1
    while f"结果{index}" in names:
        index += 1
    return f"结果{index}"


def list_sheet_names(path: str | Path) -> list[str]:
    workbook_path = Path(path)
    ensure_supported_input(workbook_path)
    suffix = workbook_path.suffix.lower()
    try:
        if suffix == ".xlsx":
            wb = openpyxl.load_workbook(
                workbook_path,
                read_only=True,
                data_only=True,
            )
            try:
                return list(wb.sheetnames)
            finally:
                wb.close()
        book = xlrd.open_workbook(str(workbook_path))
        return book.sheet_names()
    except Exception as exc:
        raise ExcelIOError(f"无法读取 Excel 文件：{exc}") from exc


def trim_trailing_empty(rows: list[list[object]]) -> list[list[object]]:
    while rows and all(value in (None, "") for value in rows[-1]):
        rows.pop()

    width = max((len(row) for row in rows), default=0)
    while width > 0:
        if any(
            len(row) >= width and row[width - 1] not in (None, "")
            for row in rows
        ):
            break
        width -= 1
    return [row[:width] for row in rows]


def read_sheet_values(path: str | Path, sheet_name: str) -> list[list[object]]:
    workbook_path = Path(path)
    ensure_supported_input(workbook_path)
    suffix = workbook_path.suffix.lower()
    try:
        if suffix == ".xlsx":
            wb = openpyxl.load_workbook(
                workbook_path,
                read_only=True,
                data_only=True,
            )
            try:
                if sheet_name not in wb.sheetnames:
                    raise ExcelIOError(f"所选 sheet 不存在：{sheet_name}")
                ws = wb[sheet_name]
                rows = [list(row) for row in ws.iter_rows(values_only=True)]
                return trim_trailing_empty(rows)
            finally:
                wb.close()

        book = xlrd.open_workbook(str(workbook_path))
        if sheet_name not in book.sheet_names():
            raise ExcelIOError(f"所选 sheet 不存在：{sheet_name}")
        sheet = book.sheet_by_name(sheet_name)
        rows = [sheet.row_values(row_index) for row_index in range(sheet.nrows)]
        return trim_trailing_empty(rows)
    except ExcelIOError:
        raise
    except Exception as exc:
        raise ExcelIOError(f"Sheet 读取失败：{exc}") from exc


def write_xlsx(path: str | Path, rows: Sequence[Sequence[object]]) -> None:
    output_path = Path(path)
    if not str(output_path):
        raise ExcelIOError("输出路径为空")
    if not output_path.parent.exists():
        raise ExcelIOError(f"输出目录不存在：{output_path.parent}")
    if output_path.suffix.lower() != ".xlsx":
        raise ExcelIOError("输出文件扩展名必须是 .xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "筛选结果"
    for row in rows:
        ws.append(list(row))

    try:
        wb.save(output_path)
    except PermissionError as exc:
        raise ExcelIOError(
            f"输出文件写入失败，可能被 Excel 或 WPS 占用：{output_path}"
        ) from exc
    except Exception as exc:
        raise ExcelIOError(f"输出文件写入失败：{exc}") from exc


def _ensure_output_path(input_path: Path, output_path: Path) -> None:
    if not str(output_path):
        raise ExcelIOError("输出路径为空")
    if output_path.suffix.lower() != ".xlsx":
        raise ExcelIOError("输出文件扩展名必须是 .xlsx")
    if not output_path.parent.exists():
        raise ExcelIOError(f"输出目录不存在：{output_path.parent}")
    if input_path.resolve() == output_path.resolve():
        raise ExcelIOError("输出路径不能与输入路径相同，请更换输出文件名")


def _append_result_sheet(wb: Workbook, result_rows: Sequence[Sequence[object]]) -> None:
    sheet_name = next_result_sheet_name(wb.sheetnames)
    ws = wb.create_sheet(sheet_name)
    for row in result_rows:
        ws.append(list(row))


def _write_result_from_xlsx(
    input_path: Path,
    output_path: Path,
    result_rows: Sequence[Sequence[object]],
) -> None:
    shutil.copy2(input_path, output_path)
    output_path.chmod(output_path.stat().st_mode | stat.S_IWRITE)
    wb = openpyxl.load_workbook(output_path, data_only=False)
    try:
        _append_result_sheet(wb, result_rows)
        wb.save(output_path)
    finally:
        wb.close()


def _write_result_from_xls(
    input_path: Path,
    output_path: Path,
    result_rows: Sequence[Sequence[object]],
) -> None:
    book = xlrd.open_workbook(str(input_path))
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    for sheet_name in book.sheet_names():
        source = book.sheet_by_name(sheet_name)
        ws = wb.create_sheet(sheet_name)
        for row_index in range(source.nrows):
            ws.append(source.row_values(row_index))
    _append_result_sheet(wb, result_rows)
    wb.save(output_path)


def write_result_workbook(
    input_path: str | Path,
    output_path: str | Path,
    result_rows: Sequence[Sequence[object]],
) -> None:
    source_path = Path(input_path)
    target_path = Path(output_path)
    ensure_supported_input(source_path)
    _ensure_output_path(source_path, target_path)
    try:
        if source_path.suffix.lower() == ".xlsx":
            _write_result_from_xlsx(source_path, target_path, result_rows)
        else:
            _write_result_from_xls(source_path, target_path, result_rows)
    except PermissionError as exc:
        raise ExcelIOError(
            f"输出文件写入失败，可能被 Excel 或 WPS 占用：{target_path}"
        ) from exc
    except ExcelIOError:
        raise
    except Exception as exc:
        raise ExcelIOError(f"输出文件写入失败：{exc}") from exc
